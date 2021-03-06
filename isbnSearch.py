'''
What the code should do:
Input ISBN
Get info from ISBN and insert into CSV File
  - Remember CSV needs pallet/lot numbers and conditions
Sync with Amazon
Host file on server
'''
#!/usr/bin/env python
import isbnlib  # For searching ISBNs
import pandas as pd  # For the dataframe to convert to .CSV
import PySimpleGUI as sg  # For the GUI
import xlrd
import openpyxl

# Functions
def tupleData(isbn, data, quan, cond, pallet, lot):
    temp = ()
    tempList = list(temp)
    for value in data.values():
        tempList.append(value)

    tempList.append(isbnlib.to_isbn10(isbn)) # Add ISBN 10
    tempList.append(quan) # Add the quantity of the book
    tempList.append(cond) # Add the condition of the book
    tempList.append(lot) # Add the lot number
    tempList.append(pallet) # Add the pallet number
    #tempList.append(isbnlib.cover(isbn))  # Add cover of ISBN

    temp = tuple(tempList)
    return temp

def get_isbn_metadata(isbn):
    # Takes the ISBN and appends information from it into a list
    # isbn.meta() gives the metadata of the ISBN in a dictionary format

    '''
    try:
        isbnMeta = isbnlib.meta(str(isbn), service='wiki') # Grabs metadata of ISBN

        return isbnMeta  # Returns dict of the ISBN meta data

    except isbnlib.ISBNLibException:
        # Try different service (Google Books) for searching the ISBN information
        isbnMeta = isbnlib.meta(str(isbn), service='goob')

        return isbnMeta
    '''
    isbnMetaList = []

    try:
        isbnMeta = isbnlib.meta(str(isbn), service='wiki')
        if bool(isbnMeta):  # Check to make sure it does not return an empty dictionary
            isbnMetaList.append(isbnMeta)
    except:
        pass
    try:
        isbnMeta = isbnlib.meta(str(isbn), service='goob')
        if bool(isbnMeta):
            isbnMetaList.append(isbnMeta)
    except:
        pass
    try:
        isbnMeta = isbnlib.meta(str(isbn), service='openl')
        if bool(isbnMeta):
            isbnMetaList.append(isbnMeta)
    except:
        pass

    layout = [
        [sg.Text("Results From Searching Wiki Books, Google Books, and OpenLibrary")],
        [sg.Text("The selection from this page will input the data for the book")],
        [sg.Listbox(values=isbnMetaList, size=(200, 6), key='-BOOK-', bind_return_key=True)],
        [sg.Button('Submit'), sg.Button('Cancel')]
    ]

    isbnMetaWindow = sg.Window('Select the Book', layout=layout)
    while True:
        eventISBNMeta, valuesISBNMeta = isbnMetaWindow.Read()
        if eventISBNMeta == sg.WIN_CLOSED or eventISBNMeta == 'Cancel':
            book = 'cancel'
            isbnMetaWindow.Close()
            return book
        if eventISBNMeta == 'Submit':
            if valuesISBNMeta['-BOOK-']:
                book = valuesISBNMeta['-BOOK-']
                isbnMetaWindow.Close()
                return book

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, header=False, index=False, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def toCSVFile(filename, isbnList):
    # Convert the list into a CSV file format
    df = pd.DataFrame(isbnList)
    append_df_to_excel(filename, df)

    return sg.popup('ISBN has been added to Inventory File')

def make_window(isbn): #Layouts cannot be re-used; must create a new window with a function
    layout_NewWindow = [
        [sg.Text('ISBN: '), sg.Text(isbn)],
        [sg.Text('Enter Quantity', size=(15, 1)), sg.InputText(key='-QUAN-')],
        [sg.Text('Enter Condition', size=(15, 1)), sg.Button('New'), sg.Button('Like New'), sg.Button('Rebind')],
        [sg.Text('Enter Pallet Number', size=(15, 1)), sg.InputText('PALLET ', key='-PALLET-')],
        [sg.Text('Enter Lot Number', size=(15, 1)), sg.InputText('LOT ', key='-LOT-')],
        [sg.Submit(), sg.Button('Back'), sg.Button('Display Inventory'), sg.Exit()] ]
    return sg.Window('Enter Inventory Information', layout_NewWindow)

def make_init_window():
    layout_new_init = [
        [sg.Text('Enter ISBN', size=(15, 1)), sg.InputText(key='-ISBN-')],
        [sg.Button("Continue"), sg.Button('Edit Existing Entry'), sg.Button('Display Inventory'),sg.Button('Exit')] ]
    return sg.Window('Enter ISBN', layout_new_init)

def make_multi_window():
    layout_multi = [
        [sg.Text('ISBN Already Exists in Inventory File')],
        [sg.Text('Would you like to add it as a new entry or edit existing instance?')],
        [sg.Button('Add New Entry'), sg.Button('Edit Existing Entry'), sg.Button('Display Inventory')],
        [sg.Exit()] ]
    return sg.Window('ISBN Already Exists', layout_multi)

def make_edit_window(toEdit):
    layout_edit = [
        [sg.Text('Locations of ISBN in Inventory:')],
        [sg.Text('Select which item to edit')],
        [sg.Listbox(values=toEdit, size=(175,6), key='-LIST-', bind_return_key=True)],
        [sg.Button('Submit'), sg.Button('Cancel')]
    ]

    editWindow = sg.Window('Edit', layout=layout_edit)
    while True:
        eventEdit, valuesEdit = editWindow.Read()
        if eventEdit == sg.WIN_CLOSED or eventEdit == 'Cancel':
            editWindow.Close()
            break
        if eventEdit == 'Submit':
            if valuesEdit['-LIST-']:
                choice = valuesEdit['-LIST-']
                update_CSV(filename, choice)
                editWindow.Close()
    return

def already_exist_check(filename, isbn):
    alreadyExist = False
    book = xlrd.open_workbook(filename)
    for sheet in book.sheets():
        for rowid in range(sheet.nrows):
            row = sheet.row(rowid)
            for colid, cell in enumerate(row):
                if cell.value == isbn:
                    alreadyExist = True
                    toEdit.append(sheet.row_values(rowid, start_colx=0))
    return alreadyExist

def find_existing_ISBN(filename, isbn):
    toEdit = []
    book = xlrd.open_workbook(filename)
    for sheet in book.sheets():
        for rowid in range(sheet.nrows):
            row = sheet.row(rowid)
            for colid, cell in enumerate(row):
                if cell.value == isbn:
                    toEdit.append(sheet.row_values(rowid, start_colx=0))
    return toEdit

def update_CSV(filename, toUpdate):
    file = openpyxl.load_workbook(filename)
    sheets = file.sheetnames
    #For reference, toUpdate[0][1] is equivalent to the title in the first list
    listEdit = []
    rowNumber = ''

    for sheet in sheets:
        currentSheet = file[sheet]

        for row in range(1, currentSheet.max_row + 1):
            isbn = False
            title = False
            quan = False
            cond = False
            lot = False
            pall = False
            for column in "ABCDEFGHIJK":

                cell_name = "{}{}".format(column, row)
                #print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                if currentSheet[cell_name].value == toUpdate[0][0]:
                    isbn = True

                if currentSheet[cell_name].value == toUpdate[0][1]:
                    title = True

                if currentSheet[cell_name].value == toUpdate[0][7]:
                    quan = True

                if currentSheet[cell_name].value == toUpdate[0][8]:
                    cond = True

                if currentSheet[cell_name].value == toUpdate[0][9]:
                    lot = True

                if currentSheet[cell_name].value == toUpdate[0][10]:
                    pall = True

            if isbn == True and title == True and quan == True and cond == True and lot == True and pall == True:
                rowNumber = cell_name[1:]

                # Set back to False since it is still looping through all instances in the file
                isbn = False
                title = False
                quan = False
                cond = False
                lot = False
                pall = False

                for col in "ABHIJK":
                    cell_name = "{}{}".format(col, rowNumber)
                    vals = currentSheet[cell_name].value
                    listEdit.append(vals)
                make_change_window(filename, listEdit, rowNumber)
    return

def make_change_window(filename, listEdit, rowNumber):
    col = [
        [sg.Text('Current')],
        [sg.Text('ISBN-13: '+listEdit[0])],
        [sg.Text('Title: '+listEdit[1])],
        [sg.Text('Quantity: '+listEdit[2])],
        [sg.Text('Condition: '+listEdit[3])],
        [sg.Text('Lot: '+listEdit[4])],
        [sg.Text('Pallet: '+listEdit[5])]
    ]

    editCol = [
        [sg.Text('Edit')],
        [sg.Text('ISBN-13: '+listEdit[0])],
        [sg.Text('Title: '+listEdit[1])],
        [sg.Text('Quantity: ', size=(9,1)), sg.InputText(listEdit[2], key='-QUAN-')],
        [sg.Text('Condition: ', size=(9,1)), sg.InputText(listEdit[3], key='-COND-')],
        [sg.Text('Lot: ', size=(9,1)), sg.InputText(listEdit[4], key='-LOT-')],
        [sg.Text('Pallet: ', size=(9,1)), sg.InputText(listEdit[5], key='-PALL-')]
    ]

    layout_change = [
        [sg.Column(col),
         sg.VSeperator(),
         sg.Column(editCol)],
        [sg.Button('Submit'), sg.Button('Cancel')]
    ]

    windowChange = sg.Window('Edit', layout_change)
    while True:
        eventChange, valuesChange = windowChange.read()
        if eventChange == sg.WIN_CLOSED or eventChange == 'Cancel':
            break
        if eventChange == 'Submit':
            file = openpyxl.load_workbook(filename)
            sheets = file.sheetnames
            for sheet in sheets:
                cs = file[sheet] #current sheet
                cs.cell(row=int(rowNumber), column=8).value = valuesChange['-QUAN-']
                cs.cell(row=int(rowNumber), column=9).value = valuesChange['-COND-']
                cs.cell(row=int(rowNumber), column=10).value = valuesChange['-LOT-']
                cs.cell(row=int(rowNumber), column=11).value = valuesChange['-PALL-']
            file.save(filename)
            sg.Popup('File Updated')
            windowChange.close()

    windowChange.close()

    return

def enterISBN(filename):
    isbnEntered = False
    initWindow = make_init_window()
    while True:
        initEvent, initValues = initWindow.read()

        while not isbnEntered:
            if initEvent == sg.WIN_CLOSED or initEvent =='Exit':
                isbn = 'Exit'
                initWindow.close()
                return isbn
            if initEvent == 'Display Inventory':
                display_CSV(filename)
                break

            if initEvent == 'Edit Existing Entry':
                toEdit = []
                isbn = initValues['-ISBN-']
                toEdit = find_existing_ISBN(filename, isbn)
                if isbn == '':
                    sg.Popup('Please Enter an ISBN')
                    break
                elif toEdit == []:
                    sg.Popup('No Existing Entries of this ISBN')
                    break
                else:
                    make_edit_window(toEdit)
                    break

            if initEvent == 'Continue':
                isbn = initValues['-ISBN-']
                if isbn == '':
                    sg.Popup('Please Enter an ISBN')
                    break
                if not isbnlib.notisbn(isbn, level='strict'):  # Verify the ISBN is valid
                    initWindow.close()
                    return isbn
                else:
                    sg.Popup("Invalid ISBN")
                    break


def display_CSV(filename):
    sg.set_options(auto_size_buttons=True)
    data = []
    header_list = []

    try:
        #df = pd.read_csv(filename, sep=',', engine='python', header=None)
        df = pd.read_excel(filename, header=None)
        header_list = df.iloc[0].tolist()
        data = df[1:].values.tolist()
    except:
        sg.popup_error('Error Reading File')
        return

    displayLayout = [
        [sg.Table(values=data, headings=header_list, display_row_numbers=True, num_rows=min(25, len(data)))]
    ]
    displayWindow = sg.Window('Inventory', displayLayout, grab_anywhere=False)
    disEvent, disValues = displayWindow.read()
    displayWindow.close()

windowMultInst_active = False #Second window for checking for multiple instances of an entered ISBN
condition = '-'
filename = 'test.xlsx'

isbn = enterISBN(filename)
window = make_window(isbn)

while True:  # Loop for window to remain open
    isbnList = []
    if isbn == 'Exit':
        break

    event, values = window.read()

    if event in (sg.WIN_CLOSED, 'Exit'):  # same as event == "Exit" or event == sg.WIN_CLOSED:
        window.close()
        break

    if event == 'Back':
        window.close()
        isbn = enterISBN(filename)
        window = make_window(isbn)

    if event == 'Display Inventory':
        display_CSV(filename)

    # If else statements for the condition buttons
    # Button colors changes based off of which condition is selected
    if event == 'New':
        condition = 'New'
        window[event].update(button_color=('white','black'))
        window['Rebind'].update(button_color=('white', 'dark blue'))
        window['Like New'].update(button_color=('white', 'dark blue'))
    if event == 'Like New':
        condition = 'Like New'
        window[event].update(button_color=('white', 'black'))
        window['Rebind'].update(button_color=('white', 'dark blue'))
        window['New'].update(button_color=('white', 'dark blue'))
    if event == 'Rebind':
        condition = 'Rebind'
        window[event].update(button_color=('white', 'black'))
        window['Like New'].update(button_color=('white', 'dark blue'))
        window['New'].update(button_color=('white', 'dark blue'))


    if event == 'Submit':
        pallet = values['-PALLET-']
        lot = values['-LOT-']
        quantity = values['-QUAN-']

        # If nothing is entered in for the pallet or lot number, set the value to a dash
        if pallet == 'PALLET ':
            pallet = '-'
        if lot == 'LOT ':
            lot = '-'
        if quantity == '':
            quantity = '-'

        if not isbnlib.notisbn(isbn, level='strict'): #Verify the ISBN is valid
            selection = True
            metadataInfo = get_isbn_metadata(isbn)
            if metadataInfo == 'cancel':
                selection = False
            if selection:
                tupleToList = tupleData(isbn, metadataInfo[0], quantity, condition, pallet, lot)
                isbnList.append(tupleToList)
                toEdit = []
                alreadyExist = already_exist_check(filename, isbn)
                toEdit = find_existing_ISBN(filename, isbn)

                if alreadyExist:
                    windowMultInst_active = True
                    window.close()
                    windowMultInst = make_multi_window()
                    while True:
                        event2, values2 = windowMultInst.Read()
                        if event2 == 'Add New Entry':
                            toCSVFile(filename, isbnList)
                            windowMultInst.Close()
                            windowMultInst_active = False
                            window = make_window(isbn)
                            condition = ' '
                            break

                        if event2 == 'Edit Existing Entry':
                            make_edit_window(toEdit)
                            windowMultInst.Close()
                            windowMultInst_active = False
                            window = make_window(isbn)
                            condition = ' '

                        if event2 == 'Display Inventory':
                            display_CSV(filename)

                        if event2 == sg.WIN_CLOSED or event2 == 'Exit':
                            windowMultInst.Close()
                            windowMultInst_active = False
                            window = make_window(isbn)
                            condition = ' '
                            break
                else:
                    toCSVFile(filename, isbnList) #Convert into the CSV File
                    window.close()
                    window = make_window(isbn)
                    condition = ' '
        else:
            sg.Popup('Not a Valid ISBN')

'''
    # Try different service (OpenL) for searching the ISBN information
    isbnList.append(isbnlib.meta(str(isbn), service='openl'))
    isbnList.append(isbnlib.cover(str(isbn)))
    return isbnList


#Code For Testing Below
Example 10 digit: 039591082X
isbn = 9780618250578
isbn = 9781680450262

'''
